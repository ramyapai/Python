from functools import reduce
from openpyxl import Workbook
import openpyxl

# Function to add elements of an erray
def addElementsofArray(arr):
    res= reduce(lambda x, y: x + y, arr)
    sum =0
    for digit in str(res):
        sum += int(digit)
    return(sum)    
            
# Function to get even indexed numbers of a string in array considering the index starts from 0
def getEvenIndexArray(n):
    evenIndexedNumbers = []    
    for i in range(len(n)):
        if i %2 != 0:
            evenIndexedNumbers.append(n[i])            
    return evenIndexedNumbers

# Function to get odd indexed numbers of a string in array considering the index starts from 0
def getOddIndexArray(n):
    oddIndexedNumbers = []    
    for i in range(len(n)):
        if i %2 == 0:
            oddIndexedNumbers.append(n[i])           
    return oddIndexedNumbers

# Function to multiply the number by 3
def multiplyThree(num):
    res = num * 3
    return res

# Function to get check digit by using the given logic
def getCheckDigit(num):
    res = num % 10
    if res == 0:
        return 0
    else:
        return 10 - res

          
# Function to calculate check digit and return the UPC-A code
def calculate_check_digit(n):
    #n = input("Please enter 12 digit string: ")
    
    oddIndexNumbers = getOddIndexArray(n)
    sumofOddIndexNum = addElementsofArray(oddIndexNumbers)
    multiplyByThree = multiplyThree(sumofOddIndexNum)
    evenIndexNumbers = getEvenIndexArray(n)
    sumofEvenIndexNum = addElementsofArray(evenIndexNumbers)
    addResult = multiplyByThree + sumofEvenIndexNum
    checkDigit = getCheckDigit(addResult)
    codeUPC = n + str(checkDigit)
    return codeUPC
    #print("check digit is {checkDigit} and UPC-A is {codeUPC}".format(checkDigit=checkDigit, codeUPC=codeUPC))
 
 # Function to read the input from excel file ("inputData.xlsx")
def readInputfromExcel():
    dataInput = openpyxl.load_workbook("inputData.xlsx")
    dataInput1 = dataInput.active
    for row in range(1, dataInput1.max_row):
        for col in dataInput1.iter_cols(1, 1):
            n = col[row].value
            codeUPC = calculate_check_digit(n) 
            
    c = dataInput1['B2']
    c.value =  codeUPC      
    dataInput.save("inputData.xlsx")        
 
 #Driver code       
readInputfromExcel()   


         
